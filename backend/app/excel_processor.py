from __future__ import annotations

import io
import re
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .models import QuoteItem, QuoteTerms, VendorQuote


class EnhancedExcelProcessor:
    """
    Parse messy Excel quotes into a normalized VendorQuote using openpyxl.

    Capabilities:
    - Handles merged cells by forward-filling values across the merged range
    - Infers header row by scanning for likely header labels
    - Maps vendor-specific columns to canonical fields (sku, description, quantity, unit_price, total, delivery)
    - Cleans numeric fields and derives totals if missing
    - Attempts vendor name extraction from top rows or filename
    """

    # Canonical column name candidates
    HEADER_CANDIDATES: Dict[str, List[str]] = {
        "sku": ["sku", "item code", "item id", "product code", "part no", "part number", "code"],
        "description": ["description", "item", "product", "material", "item description", "details"],
        "quantity": ["qty", "quantity", "qnty", "units", "pieces", "pcs"],
        "unit_price": ["unit price", "price", "rate", "unit cost", "cost", "price/ unit", "price per unit"],
        "total": ["amount", "total", "line total", "extended price", "ext price", "value"],
        "delivery": ["delivery", "lead time", "delivery time", "eta", "ship time"],
    }

    def parse(self, file_content: bytes, filename: Optional[str] = None) -> Optional[VendorQuote]:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

        best_sheet_result: Optional[Tuple[VendorQuote, int]] = None  # (quote, num_items)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            self._forward_fill_merged_cells(sheet)
            header_row_idx, header_map = self._infer_header_map(sheet)
            if header_row_idx is None or not header_map:
                continue

            items = self._extract_items(sheet, header_row_idx + 1, header_map)
            if not items:
                continue

            vendor_name = self._extract_vendor_name(sheet) or self._extract_vendor_from_filename(filename) or "Unknown Vendor"
            terms = QuoteTerms(payment="TBD", warranty="TBD")
            quote = VendorQuote(vendorName=vendor_name, items=items, terms=terms)

            if best_sheet_result is None or len(items) > best_sheet_result[1]:
                best_sheet_result = (quote, len(items))

        return best_sheet_result[0] if best_sheet_result else None

    def _forward_fill_merged_cells(self, sheet: Worksheet) -> None:
        # openpyxl stores merged cell ranges; copy top-left value to entire range
        for merged_range in list(sheet.merged_cells.ranges):
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            value = sheet.cell(row=min_row, column=min_col).value
            # Unmerge to make subsequent reads simpler
            sheet.unmerge_cells(str(merged_range))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if sheet.cell(row=r, column=c).value in (None, ""):
                        sheet.cell(row=r, column=c).value = value

    def _infer_header_map(self, sheet: Worksheet) -> Tuple[Optional[int], Dict[int, str]]:
        """
        Returns (header_row_index_zero_based, map of column_index_one_based -> canonical_name)
        """
        best_row_idx: Optional[int] = None
        best_map: Dict[int, str] = {}

        # Scan top N rows to find a likely header row
        for row_idx in range(1, min(30, sheet.max_row) + 1):
            row_values = [self._to_str(sheet.cell(row=row_idx, column=c).value) for c in range(1, sheet.max_column + 1)]
            candidate_map: Dict[int, str] = {}
            for col_idx, raw in enumerate(row_values, start=1):
                if not raw:
                    continue
                label = raw.lower().strip()
                for canonical, candidates in self.HEADER_CANDIDATES.items():
                    if any(self._fuzzy_match(label, cand) for cand in candidates):
                        if canonical not in candidate_map.values():
                            candidate_map[col_idx] = canonical
                        break
            # Consider a row a good header if it has at least description + price or quantity
            if {"description", "unit_price"}.issubset(set(candidate_map.values())) or \
               {"description", "quantity"}.issubset(set(candidate_map.values())):
                if len(candidate_map) > len(best_map):
                    best_row_idx = row_idx - 1  # zero-based for our internal use
                    best_map = candidate_map

        return best_row_idx, best_map

    def _extract_items(self, sheet: Worksheet, start_row_one_based: int, header_map: Dict[int, str]) -> List[QuoteItem]:
        items: List[QuoteItem] = []
        empty_row_streak = 0
        for row in range(start_row_one_based, sheet.max_row + 1):
            raw_cells = {canonical: self._to_str(sheet.cell(row=row, column=col_idx).value) for col_idx, canonical in header_map.items()}
            # End when we see several empty rows
            if all((v == "" for v in raw_cells.values())):
                empty_row_streak += 1
                if empty_row_streak >= 3:
                    break
                continue
            empty_row_streak = 0

            description = raw_cells.get("description", "").strip()
            if not description:
                continue
            sku = raw_cells.get("sku", "").strip() or "-"
            quantity = self._to_int(raw_cells.get("quantity")) or 1
            unit_price = self._to_float(raw_cells.get("unit_price"))
            line_total = self._to_float(raw_cells.get("total"))
            if line_total is None and unit_price is not None:
                line_total = round(unit_price * quantity, 2)
            if unit_price is None and line_total is not None and quantity:
                unit_price = round(line_total / max(quantity, 1), 4)
            delivery = raw_cells.get("delivery", "TBD") or "TBD"

            # Filter noise rows
            if not any([unit_price, line_total]):
                # If no price at all, skip
                continue

            items.append(
                QuoteItem(
                    sku=sku,
                    description=description,
                    quantity=int(quantity or 1),
                    unitPrice=float(unit_price or 0.0),
                    deliveryTime=str(delivery),
                    total=float(line_total or (unit_price or 0.0) * (quantity or 1)),
                )
            )
        return items

    def _extract_vendor_name(self, sheet: Worksheet) -> Optional[str]:
        # Search first 10 rows for vendor name hints
        top_texts: List[str] = []
        for r in range(1, min(10, sheet.max_row) + 1):
            row_vals = [self._to_str(sheet.cell(row=r, column=c).value) for c in range(1, min(10, sheet.max_column) + 1)]
            line = " ".join([v for v in row_vals if v])
            if line:
                top_texts.append(line)
        blob = "\n".join(top_texts).lower()
        # Heuristics
        m = re.search(r"vendor[:\-]\s*([a-z0-9 &().,'/-]{2,})", blob)
        if m:
            return m.group(1).strip().title()
        m = re.search(r"quote from\s*([a-z0-9 &().,'/-]{2,})", blob)
        if m:
            return m.group(1).strip().title()
        return None

    def _extract_vendor_from_filename(self, filename: Optional[str]) -> Optional[str]:
        if not filename:
            return None
        name = filename.rsplit('/', 1)[-1].rsplit('.', 1)[0]
        name = name.replace('_', ' ').replace('-', ' ')
        # Remove generic parts
        name = re.sub(r"\b(quote|quotation|vendor|price|rfq|xlsx|xls)\b", "", name, flags=re.IGNORECASE)
        name = re.sub(r"\s+", " ", name).strip()
        return name.title() if name else None

    @staticmethod
    def _to_str(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _to_float(value) -> Optional[float]:
        if value is None:
            return None
        s = str(value).strip()
        if s == "":
            return None
        s = s.replace(",", "")
        m = re.search(r"([-+]?\d*\.?\d+)", s)
        try:
            return float(m.group(1)) if m else None
        except Exception:
            return None

    @staticmethod
    def _to_int(value) -> Optional[int]:
        f = EnhancedExcelProcessor._to_float(value)
        try:
            return int(round(f)) if f is not None else None
        except Exception:
            return None

    @staticmethod
    def _fuzzy_match(label: str, candidate: str) -> bool:
        label = label.lower()
        candidate = candidate.lower()
        return candidate in label or label in candidate


enhanced_excel_processor = EnhancedExcelProcessor() 