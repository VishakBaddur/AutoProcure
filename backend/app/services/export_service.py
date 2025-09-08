import json
from typing import List, Dict, Any, Optional
from datetime import datetime
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..models import VendorQuote, MultiVendorAnalysis, MathCorrection

class ExportService:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom styles for PDF export"""
        # Title style
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        # Section header style
        self.section_style = ParagraphStyle(
            'SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.darkblue
        )
        
        # Subsection style
        self.subsection_style = ParagraphStyle(
            'Subsection',
            parent=self.styles['Heading3'],
            fontSize=14,
            spaceAfter=8,
            spaceBefore=12,
            textColor=colors.darkgreen
        )
        
        # Body text style
        self.body_style = ParagraphStyle(
            'BodyText',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
        
        # Warning style
        self.warning_style = ParagraphStyle(
            'Warning',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.red,
            spaceAfter=6
        )
        
        # Success style
        self.success_style = ParagraphStyle(
            'Success',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.green,
            spaceAfter=6
        )
    
    def export_to_pdf(self, 
                     rfq_data: Dict[str, Any],
                     analysis_result: MultiVendorAnalysis,
                     issues_detected: List[Dict[str, Any]] = None,
                     compliance_results: Dict[str, Any] = None) -> bytes:
        """Export analysis results to PDF"""
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, 
                              topMargin=72, bottomMargin=18)
        
        # Build the PDF content
        story = []
        
        # Title page
        story.append(Paragraph("AutoProcure Analysis Report", self.title_style))
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"RFQ: {rfq_data.get('title', 'N/A')}", self.section_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", self.body_style))
        story.append(PageBreak())
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", self.section_style))
        story.append(Paragraph(f"RFQ Title: {rfq_data.get('title', 'N/A')}", self.body_style))
        story.append(Paragraph(f"Description: {rfq_data.get('description', 'N/A')}", self.body_style))
        story.append(Paragraph(f"Deadline: {rfq_data.get('deadline', 'N/A')}", self.body_style))
        story.append(Paragraph(f"Total Vendors: {len(analysis_result.quotes)}", self.body_style))
        
        # Cost Summary
        if analysis_result.comparison:
            total_costs = [sum(item.total for item in quote.items) for quote in analysis_result.quotes]
            min_cost = min(total_costs)
            max_cost = max(total_costs)
            savings = max_cost - min_cost
            
            story.append(Spacer(1, 12))
            story.append(Paragraph("Cost Analysis", self.subsection_style))
            story.append(Paragraph(f"Lowest Total Cost: ${min_cost:,.2f}", self.success_style))
            story.append(Paragraph(f"Highest Total Cost: ${max_cost:,.2f}", self.warning_style))
            story.append(Paragraph(f"Potential Savings: ${savings:,.2f} ({savings/max_cost*100:.1f}%)", self.success_style))
        
        story.append(PageBreak())
        
        # Vendor Comparison Table
        story.append(Paragraph("Vendor Comparison", self.section_style))
        comparison_table = self._create_comparison_table(analysis_result.quotes)
        story.append(comparison_table)
        story.append(PageBreak())
        
        # AI Recommendation
        if analysis_result.vendor_recommendations:
            story.append(Paragraph("AI Recommendation", self.section_style))
            for rec in analysis_result.vendor_recommendations:
                if rec.is_winner:
                    story.append(Paragraph(f"🏆 WINNER: {rec.vendor_name}", self.success_style))
                    story.append(Paragraph(f"Total Cost: ${rec.total_cost:,.2f}", self.body_style))
                    story.append(Paragraph(f"Reasoning: {rec.recommendation_reason}", self.body_style))
                    if rec.items_to_purchase:
                        story.append(Paragraph("Items to Purchase:", self.subsection_style))
                        for item in rec.items_to_purchase:
                            story.append(Paragraph(f"• {item}", self.body_style))
                    story.append(Spacer(1, 12))
        
        # Issues Detected
        if issues_detected:
            story.append(Paragraph("Issues Detected", self.section_style))
            for issue in issues_detected:
                story.append(Paragraph(f"⚠️ {issue.get('type', 'Issue')}: {issue.get('description', 'N/A')}", self.warning_style))
                if issue.get('details'):
                    story.append(Paragraph(f"Details: {issue.get('details')}", self.body_style))
                story.append(Spacer(1, 6))
        
        # Compliance Results
        if compliance_results:
            story.append(Paragraph("Compliance Results", self.section_style))
            for rule, result in compliance_results.items():
                status = "✅ PASS" if result.get('passed', False) else "❌ FAIL"
                story.append(Paragraph(f"{status} {rule}: {result.get('message', 'N/A')}", 
                                     self.success_style if result.get('passed', False) else self.warning_style))
        
        # Footer
        story.append(Spacer(1, 20))
        story.append(Paragraph("Generated by AutoProcure - Intelligent Procurement Analysis", 
                              ParagraphStyle('Footer', parent=self.styles['Normal'], 
                                           fontSize=8, alignment=TA_CENTER, textColor=colors.grey)))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_comparison_table(self, quotes: List[VendorQuote]) -> Table:
        """Create vendor comparison table"""
        # Get all unique items across all quotes
        all_items = {}
        for quote in quotes:
            for item in quote.items:
                key = item.description
                if key not in all_items:
                    all_items[key] = {}
                all_items[key][quote.vendorName] = {
                    'quantity': item.quantity,
                    'unit_price': item.unitPrice,
                    'total': item.total
                }
        
        # Create table data
        table_data = [['Item', 'Vendor', 'Qty', 'Unit Price', 'Total', 'Winner']]
        
        for item_desc, vendor_data in all_items.items():
            # Find the best price for this item
            best_vendor = min(vendor_data.keys(), 
                            key=lambda v: vendor_data[v]['total'])
            
            for vendor_name, item_data in vendor_data.items():
                winner_mark = "🏆" if vendor_name == best_vendor else ""
                table_data.append([
                    item_desc,
                    vendor_name,
                    str(item_data['quantity']),
                    f"${item_data['unit_price']:.2f}",
                    f"${item_data['total']:.2f}",
                    winner_mark
                ])
        
        # Create table
        table = Table(table_data, colWidths=[2*inch, 1.5*inch, 0.5*inch, 1*inch, 1*inch, 0.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        return table
    
    def export_to_excel(self, 
                       rfq_data: Dict[str, Any],
                       analysis_result: MultiVendorAnalysis,
                       issues_detected: List[Dict[str, Any]] = None,
                       compliance_results: Dict[str, Any] = None) -> bytes:
        """Export analysis results to Excel"""
        
        buffer = io.BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        summary_ws = wb.create_sheet("Executive Summary")
        self._create_summary_sheet(summary_ws, rfq_data, analysis_result)
        
        # Comparison sheet
        comparison_ws = wb.create_sheet("Vendor Comparison")
        self._create_comparison_sheet(comparison_ws, analysis_result.quotes)
        
        # Issues sheet
        if issues_detected:
            issues_ws = wb.create_sheet("Issues Detected")
            self._create_issues_sheet(issues_ws, issues_detected)
        
        # Compliance sheet
        if compliance_results:
            compliance_ws = wb.create_sheet("Compliance Results")
            self._create_compliance_sheet(compliance_ws, compliance_results)
        
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_summary_sheet(self, ws, rfq_data: Dict[str, Any], analysis_result: MultiVendorAnalysis):
        """Create executive summary sheet"""
        # Header
        ws['A1'] = "AutoProcure Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        
        # RFQ Details
        row = 3
        ws[f'A{row}'] = "RFQ Details"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        ws[f'A{row}'] = "Title:"
        ws[f'B{row}'] = rfq_data.get('title', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "Description:"
        ws[f'B{row}'] = rfq_data.get('description', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "Deadline:"
        ws[f'B{row}'] = rfq_data.get('deadline', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "Total Vendors:"
        ws[f'B{row}'] = len(analysis_result.quotes)
        row += 2
        
        # Cost Analysis
        ws[f'A{row}'] = "Cost Analysis"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        if analysis_result.comparison:
            total_costs = [sum(item.total for item in quote.items) for quote in analysis_result.quotes]
            min_cost = min(total_costs)
            max_cost = max(total_costs)
            savings = max_cost - min_cost
            
            ws[f'A{row}'] = "Lowest Total Cost:"
            ws[f'B{row}'] = f"${min_cost:,.2f}"
            ws[f'B{row}'].font = Font(color="008000", bold=True)
            row += 1
            
            ws[f'A{row}'] = "Highest Total Cost:"
            ws[f'B{row}'] = f"${max_cost:,.2f}"
            ws[f'B{row}'].font = Font(color="FF0000", bold=True)
            row += 1
            
            ws[f'A{row}'] = "Potential Savings:"
            ws[f'B{row}'] = f"${savings:,.2f} ({savings/max_cost*100:.1f}%)"
            ws[f'B{row}'].font = Font(color="008000", bold=True)
    
    def _create_comparison_sheet(self, ws, quotes: List[VendorQuote]):
        """Create vendor comparison sheet"""
        # Get all unique items
        all_items = {}
        for quote in quotes:
            for item in quote.items:
                key = item.description
                if key not in all_items:
                    all_items[key] = {}
                all_items[key][quote.vendorName] = {
                    'quantity': item.quantity,
                    'unit_price': item.unitPrice,
                    'total': item.total
                }
        
        # Headers
        headers = ['Item', 'Vendor', 'Quantity', 'Unit Price', 'Total', 'Winner']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        row = 2
        for item_desc, vendor_data in all_items.items():
            best_vendor = min(vendor_data.keys(), 
                            key=lambda v: vendor_data[v]['total'])
            
            for vendor_name, item_data in vendor_data.items():
                ws.cell(row=row, column=1, value=item_desc)
                ws.cell(row=row, column=2, value=vendor_name)
                ws.cell(row=row, column=3, value=item_data['quantity'])
                ws.cell(row=row, column=4, value=item_data['unit_price'])
                ws.cell(row=row, column=5, value=item_data['total'])
                
                if vendor_name == best_vendor:
                    ws.cell(row=row, column=6, value="🏆 WINNER")
                    ws.cell(row=row, column=6).font = Font(color="008000", bold=True)
                
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_issues_sheet(self, ws, issues_detected: List[Dict[str, Any]]):
        """Create issues detected sheet"""
        headers = ['Type', 'Description', 'Details', 'Severity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row = 2
        for issue in issues_detected:
            ws.cell(row=row, column=1, value=issue.get('type', 'Issue'))
            ws.cell(row=row, column=2, value=issue.get('description', 'N/A'))
            ws.cell(row=row, column=3, value=str(issue.get('details', 'N/A')))
            ws.cell(row=row, column=4, value=issue.get('severity', 'Medium'))
            row += 1
    
    def _create_compliance_sheet(self, ws, compliance_results: Dict[str, Any]):
        """Create compliance results sheet"""
        headers = ['Rule', 'Status', 'Message', 'Details']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        row = 2
        for rule, result in compliance_results.items():
            status = "PASS" if result.get('passed', False) else "FAIL"
            ws.cell(row=row, column=1, value=rule)
            ws.cell(row=row, column=2, value=status)
            ws.cell(row=row, column=3, value=result.get('message', 'N/A'))
            ws.cell(row=row, column=4, value=str(result.get('details', 'N/A')))
            
            # Color code the status
            if result.get('passed', False):
                ws.cell(row=row, column=2).font = Font(color="008000", bold=True)
            else:
                ws.cell(row=row, column=2).font = Font(color="FF0000", bold=True)
            
            row += 1

# Global export service instance
export_service = ExportService()
