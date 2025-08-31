import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
import pdfplumber
import openpyxl
import io
import json
from typing import List, Dict, Any, Optional
import httpx
from .models import QuoteItem, QuoteTerms, VendorQuote, AnalysisResult, MultiVendorAnalysis
from .slack import send_slack_alert
from .ai_processor import ai_processor
from .multi_vendor_analyzer import multi_vendor_analyzer
from .database import db
from .pdf_processor import enhanced_pdf_processor
from .excel_processor import enhanced_excel_processor
# Import new analysis modules
from .obfuscation_detector import obfuscation_detector
from .math_validator import math_validator
from .justification_helper import justification_helper
from .delay_tracker import delay_tracker
from .currency_handler import currency_handler

app = FastAPI(title="AutoProcure API", version="1.0.0")

# CORS middleware for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000", 
        "http://127.0.0.1:3000",
        "https://auto-procure.vercel.app",
        "https://autoprocure-ai.vercel.app",
        "https://autoprocure-procurement.vercel.app",
        "https://autoprocure-frontend.onrender.com",
        "https://autoprocure.onrender.com"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Remove authentication dependencies for beta
# Remove: from .auth import auth_manager
# Remove: security = HTTPBearer()
# Remove: get_current_user function
# Remove: SignupRequest, LoginRequest, /auth/signup, /auth/login, /auth/me endpoints
# Remove: current_user and Depends(get_current_user) from all endpoints

# Waitlist models
class WaitlistRequest(BaseModel):
    email: str

class WaitlistResponse(BaseModel):
    success: bool
    message: str
    id: Optional[str] = None

def extract_text_from_pdf(file_content: bytes) -> str:
    """Extract text from PDF using enhanced processor with OCR fallback"""
    try:
        # Save file content to temporary file
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            tmp_file.write(file_content)
            tmp_file_path = tmp_file.name
        
        # Use enhanced PDF processor
        result = enhanced_pdf_processor.extract_text_enhanced(tmp_file_path)
        
        # Clean up temporary file
        import os
        os.unlink(tmp_file_path)
        
        # Log extraction method used
        print(f"[PDF EXTRACTION] Method: {result['extraction_method']}, Success: {result['success']}")
        
        return result['text']
    except Exception as e:
        print(f"[PDF EXTRACTION ERROR] {str(e)}")
        # Fallback to original method
        try:
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                text = ""
                for page_num, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    print(f"=== Extracted PDF Page {page_num+1} Text ===")
                    print(page_text)
                    text += page_text
                return text
        except Exception as fallback_error:
            print(f"[PDF FALLBACK ERROR] {str(fallback_error)}")
            return f"PDF extraction failed: {str(e)}"

def extract_text_from_excel(file_content: bytes) -> str:
    """Extract text from Excel using openpyxl"""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                text += " ".join(str(cell) for cell in row if cell) + "\n"
        return text
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel parsing error: {str(e)}")

def parse_csv_to_quote(file_content: bytes, filename: str) -> VendorQuote:
    """Parse CSV file directly into a VendorQuote object"""
    try:
        import csv
        csv_content = file_content.decode('utf-8')
        csv_reader = csv.reader(io.StringIO(csv_content))
        
        # Skip header row
        header = next(csv_reader, None)
        if not header:
            raise HTTPException(status_code=400, detail="No data found in CSV file")
        
        vendor_name = "Unknown Vendor"
        items = []
        
        for row_num, row in enumerate(csv_reader, 1):
            if len(row) >= 6:  # Ensure we have enough columns
                vendor = row[0] if row[0] else "Unknown"
                description = row[1] if row[1] else f"Item {row_num}"
                sku = row[2] if row[2] else f"SKU-{row_num}"
                quantity_str = row[3] if row[3] else "1"
                unit_price_str = row[4] if row[4] else "0"
                total_str = row[5] if row[5] else "0"
                delivery_time = row[6] if len(row) > 6 and row[6] else "N/A"
                
                # Set vendor name from first row
                if row_num == 1:
                    vendor_name = vendor
                
                # Convert strings to numbers
                try:
                    quantity = float(quantity_str) if quantity_str else 1
                    unit_price = float(unit_price_str.replace('$', '').replace(',', '')) if unit_price_str else 0
                    total = float(total_str.replace('$', '').replace(',', '')) if total_str else 0
                except ValueError:
                    quantity = 1
                    unit_price = 0
                    total = 0
                
                # Create QuoteItem
                item = QuoteItem(
                    sku=sku,
                    description=description,
                    quantity=quantity,
                    unitPrice=unit_price,
                    deliveryTime=delivery_time,
                    total=total
                )
                items.append(item)
        
        # Create VendorQuote
        quote = VendorQuote(
            vendorName=vendor_name,
            items=items,
            terms=QuoteTerms(payment="N/A", warranty="N/A"),
            reliability_score=None,
            delivery_rating=None,
            quality_rating=None
        )
        
        return quote
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"CSV parsing error: {str(e)}")

def extract_text_from_csv(file_content: bytes) -> str:
    """Extract text from CSV files and format for AI analysis"""
    try:
        import csv
        csv_content = file_content.decode('utf-8')
        csv_reader = csv.reader(io.StringIO(csv_content))
        
        # Skip header row
        header = next(csv_reader, None)
        if not header:
            return "No data found in CSV file"
        
        # Format as structured quote text
        text = f"Vendor Quote Analysis:\n"
        text += f"Vendor: {header[0] if len(header) > 0 else 'Unknown'}\n\n"
        text += "Items:\n"
        
        for row_num, row in enumerate(csv_reader, 1):
            if len(row) >= 6:  # Ensure we have enough columns
                vendor = row[0] if row[0] else "Unknown"
                description = row[1] if row[1] else "Unknown Item"
                sku = row[2] if row[2] else f"SKU-{row_num}"
                quantity = row[3] if row[3] else "1"
                unit_price = row[4] if row[4] else "0"
                total = row[5] if row[5] else "0"
                delivery_time = row[6] if len(row) > 6 and row[6] else "N/A"
                
                text += f"Item {row_num}: {description} (SKU: {sku})\n"
                text += f"  Quantity: {quantity}\n"
                text += f"  Unit Price: ${unit_price}\n"
                text += f"  Total: ${total}\n"
                text += f"  Delivery Time: {delivery_time}\n\n"
        
        return text
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"CSV parsing error: {str(e)}")

@app.on_event("startup")
async def startup_event():
    """Initialize database connection on startup"""
    try:
        await db.connect()
        print("✅ Database connected successfully")
    except Exception as e:
        print(f"⚠️ Database connection failed: {e}")
        print("⚠️ App will continue without database functionality")

@app.on_event("shutdown")
async def shutdown_event():
    """Close database connections on shutdown"""
    try:
        if db.pool:
            await db.pool.close()
            print("✅ Database connection closed")
    except Exception as e:
        print(f"⚠️ Error closing database: {e}")

# Remove get_current_user and all auth endpoints
# Remove current_user from upload_file, analyze_multiple_quotes, get_quote_history, get_quote, get_analytics

# For endpoints that used current_user, set user_id, user_email, user_name to None
# For quote history and analytics, return all data (or empty if not available)

@app.get("/")
async def root():
    return {"message": "AutoProcure API is running!"}

@app.post("/upload", response_model=AnalysisResult)
async def upload_file(
    file: UploadFile = File(...),
):
    """Upload and analyze vendor quote files with RAG context"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    file_extension = file.filename.lower().split('.')[-1]
    if file_extension not in ['pdf', 'xlsx', 'xls', 'csv']:
        raise HTTPException(status_code=400, detail="Unsupported file type. Please upload PDF, Excel, or CSV files.")
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Extract text based on file type
        if file_extension == 'pdf':
            text_content = extract_text_from_pdf(file_content)
            parsed_quote = await ai_processor.analyze_quote(text_content)
        elif file_extension == 'csv':
            # Handle CSV files - create structured quote directly
            parsed_quote = parse_csv_to_quote(file_content, file.filename)
            text_content = f"CSV Quote from {parsed_quote.vendorName}: {len(parsed_quote.items)} items"
        else:
            # Try structured Excel first
            structured_quote = enhanced_excel_processor.parse(file_content, filename=file.filename)
            if structured_quote and structured_quote.items:
                parsed_quote = structured_quote
                text_content = "\n".join([f"{it.quantity} x {it.description} @ {it.unitPrice}" for it in structured_quote.items])
            else:
                text_content = extract_text_from_excel(file_content)
                parsed_quote = await ai_processor.analyze_quote(text_content)
        
        # Get RAG context if available
        user_id = None  # Set user_id to None for public endpoints
        rag_context = ""
        if user_id and parsed_quote.items:
            skus = [item.sku for item in parsed_quote.items if item.sku]
            if skus:
                past_quotes = await db.get_relevant_past_quotes(user_id, skus, limit=5)
                if past_quotes:
                    rag_context = "\n".join([
                        f"Date: {q['created_at']}, Vendor: {q['vendor_name']}, SKU: {q['sku']}, Desc: {q['description']}, Qty: {q['quantity']}, Unit: {q['unit_price']}, Total: {q['total']}" for q in past_quotes
                    ])
        
        # If initial analysis was text-based, re-run with RAG context where applicable
        if not isinstance(parsed_quote, VendorQuote) or not parsed_quote.items:
            parsed_quote = await ai_processor.analyze_quote(text_content, rag_context=rag_context)
        
        # Debug output
        print(f"[DEBUG] Quote analysis result:")
        print(f"  Vendor: {parsed_quote.vendorName}")
        print(f"  Items count: {len(parsed_quote.items)}")
        print(f"  First item: {parsed_quote.items[0].description if parsed_quote.items else 'No items'}")
        
        # Create comparison and recommendation
        total_cost = sum(item.total for item in parsed_quote.items)
        delivery_time = parsed_quote.items[0].deliveryTime if parsed_quote.items else "N/A"
        comparison = {
            "totalCost": total_cost,
            "deliveryTime": delivery_time,
            "vendorCount": 1
        }
        
        # Generate smart recommendation based on actual analysis
        if parsed_quote.items:
            recommendation = f"✅ Quote analyzed successfully. Total cost: ${total_cost:,.2f} from {parsed_quote.vendorName}. {len(parsed_quote.items)} items identified."
        else:
            recommendation = "⚠️ Quote analysis completed but no items were found. Please verify the document format."
        
        # Run advanced analysis features
        advanced_analysis = await run_advanced_analysis([parsed_quote], [text_content])
        
        result = AnalysisResult(
            quotes=[parsed_quote],
            comparison=comparison,
            recommendation=recommendation,
            advanced_analysis=advanced_analysis
        )
        
        # Save to database
        user_email = None # Set user_email to None for public endpoints
        user_name = None # Set user_name to None for public endpoints
        
        # await auth_manager._create_user_record(user_id, user_email, user_name) # Removed auth_manager call
        
        quote_id = await db.save_quote_analysis(
            filename=file.filename,
            file_type=file_extension,
            raw_text=text_content,
            analysis_result=result,
            user_id=user_id
        )
        
        # Add quote ID to response
        result_dict = result.dict()
        result_dict["quote_id"] = quote_id
        
        # Send Slack alert
        await send_slack_alert(result)
        
        return result_dict
        
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Processing error: {str(e)}")

@app.post("/analyze-multiple", response_model=AnalysisResult)
async def analyze_multiple_quotes(
    files: List[UploadFile] = File(...),
):
    """Analyze multiple vendor quotes and provide intelligent multi-vendor recommendations"""
    if not files or len(files) < 2:
        raise HTTPException(status_code=400, detail="At least 2 vendor quotes required for comparison")
    
    if len(files) > 5:
        raise HTTPException(status_code=400, detail="Maximum 5 vendor quotes allowed for analysis")
    
    try:
        quotes = []
        file_contents = []
        raw_texts = []
        
        # Process each file
        for file in files:
            if not file.filename:
                continue
                
            file_extension = file.filename.lower().split('.')[-1]
            if file_extension not in ['pdf', 'xlsx', 'xls', 'csv']:
                continue
            
            # Read and extract text
            file_content = await file.read()
            if file_extension == 'pdf':
                text_content = extract_text_from_pdf(file_content)
                parsed_quote = await ai_processor.analyze_quote(text_content)
            elif file_extension == 'csv':
                # Handle CSV files - create structured quote directly
                parsed_quote = parse_csv_to_quote(file_content, file.filename)
                text_content = f"CSV Quote from {parsed_quote.vendorName}: {len(parsed_quote.items)} items"
            else:
                # Structured Excel first
                structured_quote = enhanced_excel_processor.parse(file_content, filename=file.filename)
                if structured_quote and structured_quote.items:
                    parsed_quote = structured_quote
                    text_content = "\n".join([f"{it.quantity} x {it.description} @ {it.unitPrice}" for it in structured_quote.items])
                else:
                    text_content = extract_text_from_excel(file_content)
                    parsed_quote = await ai_processor.analyze_quote(text_content)
            
            # Analyze quote
            quote = parsed_quote
            quotes.append(quote)
            file_contents.append({
                "filename": file.filename,
                "content": text_content
            })
            raw_texts.append(text_content)
        
        if len(quotes) < 2:
            raise HTTPException(status_code=400, detail="At least 2 valid quotes required for comparison")
        
        # Get RAG context if available
        user_id = None  # Set user_id to None for public endpoints
        rag_context = ""
        if user_id and quotes:
            all_skus = []
            for quote in quotes:
                all_skus.extend([item.sku for item in quote.items if item.sku])
            if all_skus:
                past_quotes = await db.get_relevant_past_quotes(user_id, all_skus, limit=5)
                if past_quotes:
                    rag_context = "\n".join([
                        f"Date: {q['created_at']}, Vendor: {q['vendor_name']}, SKU: {q['sku']}, Desc: {q['description']}, Qty: {q['quantity']}, Unit: {q['unit_price']}, Total: {q['total']}" for q in past_quotes
                    ])
        
        # Perform multi-vendor analysis
        multi_vendor_result = await multi_vendor_analyzer.analyze_multiple_quotes(quotes, rag_context)

        # Suggestion/Conclusion logic
        def suggest_best_vendor(quotes):
            # Build a map of best price for each item across vendors
            item_best = {}
            for quote in quotes:
                for item in quote.items:
                    key = item.description.strip().lower()
                    if key not in item_best or item.unitPrice < item_best[key]['unitPrice']:
                        item_best[key] = {
                            'vendor': quote.vendorName,
                            'unitPrice': item.unitPrice,
                            'total': item.total,
                            'quantity': item.quantity
                        }
            
            # Calculate totals for each vendor
            vendor_totals = {}
            for quote in quotes:
                vendor_totals[quote.vendorName] = sum(item.total for item in quote.items)
            
            # Find best single vendor
            best_vendor = min(vendor_totals, key=vendor_totals.get)
            best_vendor_total = vendor_totals[best_vendor]
            
            # Calculate split order savings
            split_total = sum(x['unitPrice'] * x['quantity'] for x in item_best.values())
            savings = best_vendor_total - split_total
            
            # Generate recommendation
            if savings > best_vendor_total * 0.05:  # More than 5% savings
                return f"🎯 **RECOMMENDATION**: Split your order across vendors for maximum savings! You can save ${savings:.2f} by purchasing each item from the vendor offering the lowest price, compared to buying everything from {best_vendor} (${best_vendor_total:.2f})."
            else:
                return f"🎯 **RECOMMENDATION**: Choose {best_vendor} for simplicity and reliability. Total cost: ${best_vendor_total:.2f}. The savings from splitting the order (${savings:.2f}) don't justify the additional complexity."

        suggestion = suggest_best_vendor(quotes)
        
        # Create analysis result
        total_cost = sum(
            sum(item.total for item in quote.items) 
            for quote in quotes
        )
        
        comparison = {
            "totalCost": total_cost,
            "vendorCount": len(quotes),
            "costSavings": multi_vendor_result.cost_savings,
            "riskAssessment": multi_vendor_result.risk_assessment
        }
        
        # Run advanced analysis features
        advanced_analysis = await run_advanced_analysis(quotes, raw_texts)
        
        result = AnalysisResult(
            quotes=[],  # Don't duplicate quotes here since they're in multi_vendor_analysis
            comparison=comparison,
            recommendation=suggestion,  # Use the clean suggestion instead of AI's raw output
            multi_vendor_analysis=multi_vendor_result,
            advanced_analysis=advanced_analysis
        )
        
        # Save to database
        user_email = None # Set user_email to None for public endpoints
        user_name = None # Set user_name to None for public endpoints
        
        # await auth_manager._create_user_record(user_id, user_email, user_name) # Removed auth_manager call
        
        # Save each quote separately for history
        quote_ids = []
        for i, file_content in enumerate(file_contents):
            quote_id = await db.save_quote_analysis(
                filename=file_content["filename"],
                file_type="multi_vendor",
                raw_text=file_content["content"],
                analysis_result=result,
                user_id=user_id
            )
            quote_ids.append(quote_id)
        
        # Add quote IDs and suggestion to response
        result_dict = result.dict()
        result_dict["quote_ids"] = quote_ids
        result_dict["suggestion"] = suggestion
        
        # Send Slack alert
        await send_slack_alert(result)
        
        return result_dict
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Multi-vendor analysis failed: {str(e)}")

async def run_advanced_analysis(quotes: List[VendorQuote], raw_texts: List[str]) -> Dict[str, Any]:
    """Run all advanced analysis features"""
    advanced_analysis = {}
    
    try:
        # 1. Obfuscation Detection
        obfuscation_results = []
        for i, quote in enumerate(quotes):
            raw_text = raw_texts[i] if i < len(raw_texts) else ""
            obfuscation_result = obfuscation_detector.analyze_quote(quote, raw_text)
            obfuscation_results.append({
                "vendor": quote.vendorName,
                "analysis": obfuscation_result
            })
        advanced_analysis["obfuscation_detection"] = {
            "results": obfuscation_results,
            "summary": "Obfuscation analysis completed"
        }
        
        # 2. Math Validation
        validation_results = []
        for quote in quotes:
            validation_result = math_validator.validate_quote(quote)
            validation_results.append({
                "vendor": quote.vendorName,
                "validation": validation_result
            })
        advanced_analysis["math_validation"] = {
            "results": validation_results,
            "summary": "Math validation completed"
        }
        
        # 3. Justification Helper (for multi-vendor scenarios)
        if len(quotes) > 1:
            # Find the selected vendor (lowest cost for this example)
            selected_vendor = min(quotes, key=lambda q: sum(item.total for item in q.items))
            justification_result = justification_helper.generate_justification(selected_vendor, quotes)
            advanced_analysis["justification_helper"] = {
                "selected_vendor": selected_vendor.vendorName,
                "justification": justification_result
            }
        
        # 4. Delay Tracker
        delay_result = delay_tracker.analyze_timeline_risks(quotes, raw_texts)
        advanced_analysis["delay_tracker"] = delay_result
        
    except Exception as e:
        print(f"Advanced analysis error: {str(e)}")
        advanced_analysis["error"] = f"Advanced analysis failed: {str(e)}"
    
    return advanced_analysis

@app.get("/quotes")
async def get_quote_history(
    limit: int = 10,
):
    """Get quote history for current user"""
    try:
        user_id = None # Set user_id to None for public endpoints
        quotes = await db.get_quote_history(user_id=user_id, limit=limit)
        return {"quotes": quotes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get quote history: {str(e)}")

@app.get("/quotes/{quote_id}")
async def get_quote(
    quote_id: str,
):
    """Get specific quote by ID"""
    try:
        quote = await db.get_quote_by_id(quote_id)
        if not quote:
            raise HTTPException(status_code=404, detail="Quote not found")
        
        # Check if user owns this quote (if authenticated)
        # if current_user and quote.get("user_id") and quote["user_id"] != current_user["user_id"]: # Removed auth check
        #     raise HTTPException(status_code=403, detail="Access denied")
            
        return quote
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get quote: {str(e)}")

@app.get("/analytics")
async def get_analytics(
):
    """Get analytics data for current user"""
    try:
        user_id = None # Set user_id to None for public endpoints
        analytics = await db.get_analytics(user_id=user_id)
        return analytics
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get analytics: {str(e)}")

@app.get("/health")
async def health_check():
    """Health check endpoint that works without database"""
    try:
        # Check if database is connected
        db_status = "connected" if db.pool else "disconnected"
        
        return {
            "status": "healthy", 
            "service": "AutoProcure API",
            "database": db_status,
            "timestamp": "2024-01-01T00:00:00Z"
        }
    except Exception as e:
        return {
            "status": "degraded",
            "service": "AutoProcure API", 
            "error": str(e),
            "timestamp": "2024-01-01T00:00:00Z"
        }

@app.get("/ai-status")
async def ai_status():
    """Check AI provider status"""
    ai_provider = os.getenv('AI_PROVIDER', 'ollama')
    model_name = os.getenv('AI_MODEL', 'mistral')
    
    # Test Ollama if it's the provider
    ollama_working = False
    if ai_provider == 'ollama':
        try:
            # Simple test prompt
            test_prompt = "Say 'Hello World'"
            response = await ai_processor._call_ollama(test_prompt)
            ollama_working = len(response.strip()) > 0
        except Exception as e:
            print(f"Ollama test failed: {str(e)}")
            ollama_working = False
    
    return {
        "ai_provider": ai_provider,
        "model_name": model_name,
        "ollama_url": os.getenv('OLLAMA_URL', 'http://localhost:11434'),
        "ollama_working": ollama_working,
        "openai_configured": bool(os.getenv('OPENAI_API_KEY')),
        "database_connected": db.pool is not None,
        "supabase_auth_configured": bool(os.getenv('SUPABASE_URL') and os.getenv('SUPABASE_ANON_KEY'))
    }

@app.get("/test-nlp")
async def test_nlp():
    """Test NLP analysis with sample quote text"""
    try:
        sample_text = """
        Quote from ABC Supplies Inc.
        
        ITEMS:
        10 x Laptop Computers @ $899.99
        5 x Monitors @ $299.99
        
        Payment: Net 30
        Warranty: 1 year standard warranty
        """
        
        # Test the NLP analysis directly with the text
        result = ai_processor._analyze_quote_with_nlp(sample_text)
        
        return {
            "status": "success",
            "sample_text": sample_text,
            "analysis_result": result,
            "parsed_result": json.loads(result)
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e)
        }

@app.post("/waitlist", response_model=WaitlistResponse)
async def join_waitlist(request: WaitlistRequest):
    """Add email to waitlist"""
    try:
        # Basic email validation
        if not request.email or '@' not in request.email:
            raise HTTPException(status_code=400, detail="Invalid email address")
        
        result = await db.add_to_waitlist(request.email)
        
        if result["success"]:
            return WaitlistResponse(
                success=True,
                message=result["message"],
                id=result.get("id")
            )
        else:
            return WaitlistResponse(
                success=False,
                message=result["message"]
            )
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to join waitlist: {str(e)}")

@app.get("/waitlist/count")
async def get_waitlist_count():
    """Get total number of waitlist subscribers"""
    try:
        count = await db.get_waitlist_count()
        return {"count": count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get waitlist count: {str(e)}")

# Obfuscation Detection Feedback Model
class ObfuscationFeedback(BaseModel):
    quote_id: str
    pattern_found: str
    category: str  # "hidden_fees", "bundled_pricing", "conditional_pricing", "complex_structures"
    is_correct: bool
    user_notes: Optional[str] = None
    suggested_pattern: Optional[str] = None

@app.post("/obfuscation/feedback")
async def submit_obfuscation_feedback(feedback: ObfuscationFeedback):
    """Submit feedback on obfuscation detection to improve accuracy"""
    try:
        # Log the feedback
        await db.log_obfuscation_feedback(
            quote_id=feedback.quote_id,
            pattern_found=feedback.pattern_found,
            category=feedback.category,
            is_correct=feedback.is_correct,
            user_notes=feedback.user_notes,
            suggested_pattern=feedback.suggested_pattern
        )
        
        # If user suggests a new pattern and marks current detection as incorrect
        if feedback.suggested_pattern and not feedback.is_correct:
            # Learn the new pattern with high confidence
            obfuscation_detector.learn_new_pattern(
                category=feedback.category,
                pattern=feedback.suggested_pattern,
                confidence=0.9
            )
        
        # If user confirms a detection is correct, increase confidence
        elif feedback.is_correct:
            # The pattern is already in our system, so we can increase its weight
            print(f"Confirmed obfuscation pattern: {feedback.pattern_found} in category {feedback.category}")
        
        return {
            "success": True,
            "message": "Feedback submitted successfully. Thank you for helping improve our detection accuracy!",
            "pattern_learned": bool(feedback.suggested_pattern and not feedback.is_correct)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to submit feedback: {str(e)}")

@app.get("/obfuscation/accuracy-stats")
async def get_obfuscation_accuracy_stats():
    """Get statistics on obfuscation detection accuracy"""
    try:
        # Get feedback statistics from database
        stats = await db.get_obfuscation_accuracy_stats()
        
        return {
            "success": True,
            "accuracy_stats": {
                "total_feedback": stats.get("total_feedback", 0),
                "correct_detections": stats.get("correct_detections", 0),
                "incorrect_detections": stats.get("incorrect_detections", 0),
                "accuracy_percentage": stats.get("accuracy_percentage", 0),
                "patterns_learned": stats.get("patterns_learned", 0),
                "categories": {
                    "hidden_fees": stats.get("hidden_fees_accuracy", 0),
                    "bundled_pricing": stats.get("bundled_pricing_accuracy", 0),
                    "conditional_pricing": stats.get("conditional_pricing_accuracy", 0),
                    "complex_structures": stats.get("complex_structures_accuracy", 0)
                }
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get accuracy stats: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
