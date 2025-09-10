import io
import tempfile
import os
from typing import Dict, Any, Optional
import pdfplumber
import openpyxl
import csv
from PIL import Image
import pytesseract
import fitz  # PyMuPDF
import tabula
import numpy as np

# Optional import for image processing
try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False
    print("⚠️  OpenCV not available - image processing features disabled")

class EnhancedFileProcessor:
    """Comprehensive file processor that handles any format including scanned documents"""
    
    def __init__(self):
        self.supported_formats = {
            'pdf': self.process_pdf,
            'xlsx': self.process_excel,
            'xls': self.process_excel,
            'csv': self.process_csv,
            'txt': self.process_text,
            'jpg': self.process_image,
            'jpeg': self.process_image,
            'png': self.process_image,
            'tiff': self.process_image,
            'tif': self.process_image,
            'bmp': self.process_image
        }
    
    def process_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process any file type and extract structured data"""
        try:
            file_extension = filename.lower().split('.')[-1]
            
            # Special handling for text files with .pdf extension
            if file_extension == 'pdf':
                # Try to decode as text first (for text files with .pdf extension)
                try:
                    text_content = file_content.decode('utf-8')
                    # If it's mostly readable text, treat it as a text file
                    printable_chars = sum(1 for c in text_content if c.isprintable() or c in '\n\r\t')
                    if printable_chars / len(text_content) > 0.9 and len(text_content) > 10:
                        print(f"[ENHANCED PROCESSOR] Detected text file with .pdf extension: {filename}")
                        return {
                            'success': True,
                            'text': text_content,
                            'method': 'text_as_pdf',
                            'structured_data': self._parse_text_to_structured(text_content)
                        }
                except UnicodeDecodeError:
                    # If it's not readable text, process as actual PDF
                    print(f"[ENHANCED PROCESSOR] Processing as real PDF: {filename}")
                    pass
            
            if file_extension not in self.supported_formats:
                return {
                    'success': False,
                    'error': f'Unsupported file format: {file_extension}',
                    'text': '',
                    'structured_data': None
                }
            
            processor = self.supported_formats[file_extension]
            return processor(file_content, filename)
            
        except Exception as e:
            print(f"[ENHANCED PROCESSOR ERROR] {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'text': '',
                'structured_data': None
            }
    
    def process_pdf(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process PDF files including scanned documents"""
        try:
            # First, check if this is actually a text file with .pdf extension
            try:
                text_content = file_content.decode('utf-8')
                # If it's mostly readable text, treat it as a text file
                printable_chars = sum(1 for c in text_content if c.isprintable() or c in '\n\r\t')
                if printable_chars / len(text_content) > 0.9 and len(text_content) > 10:
                    print(f"[PDF PROCESSOR] Detected text file with .pdf extension: {filename}")
                    return {
                        'success': True,
                        'text': text_content,
                        'method': 'text_as_pdf',
                        'structured_data': self._parse_text_to_structured(text_content)
                    }
            except UnicodeDecodeError:
                # Not a text file, proceed with PDF processing
                pass
            
            # Method 1: Try pdfplumber for text-based PDFs
            text = self._extract_text_pdfplumber(file_content)
            if text.strip():
                return {
                    'success': True,
                    'text': text,
                    'method': 'pdfplumber',
                    'structured_data': self._parse_text_to_structured(text)
                }
            
            # Method 2: Try PyMuPDF for better text extraction
            text = self._extract_text_pymupdf(file_content)
            if text.strip():
                return {
                    'success': True,
                    'text': text,
                    'method': 'pymupdf',
                    'structured_data': self._parse_text_to_structured(text)
                }
            
            # Method 3: OCR for scanned documents
            text = self._extract_text_ocr(file_content)
            if text.strip():
                return {
                    'success': True,
                    'text': text,
                    'method': 'ocr',
                    'structured_data': self._parse_text_to_structured(text)
                }
            
            # Method 4: Table extraction
            text = self._extract_tables(file_content)
            if text.strip():
                return {
                    'success': True,
                    'text': text,
                    'method': 'table_extraction',
                    'structured_data': self._parse_text_to_structured(text)
                }
            
            return {
                'success': False,
                'error': 'No text could be extracted from PDF',
                'text': '',
                'structured_data': None
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'PDF processing failed: {str(e)}',
                'text': '',
                'structured_data': None
            }
    
    def _extract_text_pdfplumber(self, file_content: bytes) -> str:
        """Extract text using pdfplumber"""
        try:
            with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"
                return text
        except Exception as e:
            print(f"pdfplumber failed: {e}")
            return ""
    
    def _extract_text_pymupdf(self, file_content: bytes) -> str:
        """Extract text using PyMuPDF"""
        try:
            doc = fitz.open(stream=file_content, filetype="pdf")
            text = ""
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                page_text = page.get_text()
                text += page_text + "\n"
            doc.close()
            return text
        except Exception as e:
            print(f"PyMuPDF failed: {e}")
            return ""
    
    def _extract_text_ocr(self, file_content: bytes) -> str:
        """Extract text using OCR for scanned documents"""
        try:
            doc = fitz.open(stream=file_content, filetype="pdf")
            text = ""
            
            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                
                # Convert page to high-resolution image
                mat = fitz.Matrix(3, 3)  # 3x zoom for better OCR
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                
                # Convert to PIL Image
                img = Image.open(io.BytesIO(img_data))
                
                # Preprocess image for better OCR
                img = self._preprocess_image_for_ocr(img)
                
                # OCR the image
                page_text = pytesseract.image_to_string(img, config='--psm 6')
                text += page_text + "\n"
            
            doc.close()
            return text
            
        except Exception as e:
            print(f"OCR failed: {e}")
            return ""
    
    def _preprocess_image_for_ocr(self, img: Image.Image) -> Image.Image:
        """Preprocess image for better OCR results"""
        try:
            if not CV2_AVAILABLE:
                # If OpenCV is not available, just convert to grayscale
                if img.mode != 'L':
                    img = img.convert('L')
                return img
            
            # Convert to numpy array
            img_array = np.array(img)
            
            # Convert to grayscale if needed
            if len(img_array.shape) == 3:
                gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
            else:
                gray = img_array
            
            # Apply thresholding to get black text on white background
            _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
            
            # Denoise
            denoised = cv2.medianBlur(thresh, 3)
            
            # Convert back to PIL Image
            return Image.fromarray(denoised)
            
        except Exception as e:
            print(f"Image preprocessing failed: {e}")
            return img
    
    def _extract_tables(self, file_content: bytes) -> str:
        """Extract tables from PDF"""
        try:
            tables = tabula.read_pdf(io.BytesIO(file_content), pages='all')
            text = ""
            for i, table in enumerate(tables):
                text += f"Table {i+1}:\n{table.to_string()}\n\n"
            return text
        except Exception as e:
            print(f"Table extraction failed: {e}")
            return ""
    
    def process_excel(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process Excel files"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            text = ""
            structured_data = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    text += " ".join(row_data) + "\n"
                    sheet_data.append(row_data)
                
                structured_data.append({
                    'sheet_name': sheet_name,
                    'data': sheet_data
                })
            
            return {
                'success': True,
                'text': text,
                'method': 'excel',
                'structured_data': structured_data
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Excel processing failed: {str(e)}',
                'text': '',
                'structured_data': None
            }
    
    def process_csv(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process CSV files"""
        try:
            csv_content = file_content.decode('utf-8')
            csv_reader = csv.reader(io.StringIO(csv_content))
            
            text = ""
            structured_data = []
            
            for row in csv_reader:
                row_data = [str(cell) if cell else "" for cell in row]
                text += " ".join(row_data) + "\n"
                structured_data.append(row_data)
            
            return {
                'success': True,
                'text': text,
                'method': 'csv',
                'structured_data': structured_data
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'CSV processing failed: {str(e)}',
                'text': '',
                'structured_data': None
            }
    
    def process_text(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process text files"""
        try:
            # Decode text content
            text = file_content.decode('utf-8')
            
            return {
                'success': True,
                'text': text,
                'method': 'text',
                'structured_data': self._parse_text_to_structured(text)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Text processing failed: {str(e)}',
                'text': '',
                'structured_data': None
            }
    
    def process_image(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """Process image files using OCR"""
        try:
            # Load image
            img = Image.open(io.BytesIO(file_content))
            
            # Preprocess image for better OCR
            img = self._preprocess_image_for_ocr(img)
            
            # Extract text using OCR
            if not CV2_AVAILABLE:
                print("⚠️  Using basic OCR without image preprocessing")
            text = pytesseract.image_to_string(img, config='--psm 6')
            
            return {
                'success': True,
                'text': text,
                'method': 'image_ocr',
                'structured_data': self._parse_text_to_structured(text)
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Image processing failed: {str(e)}',
                'text': '',
                'structured_data': None
            }
    
    def _parse_text_to_structured(self, text: str) -> Dict[str, Any]:
        """Parse extracted text into structured data"""
        try:
            # Simple parsing logic - can be enhanced with AI
            lines = text.split('\n')
            vendor_name = "Unknown Vendor"
            items = []
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Try to extract vendor name
                if any(keyword in line.lower() for keyword in ['vendor', 'supplier', 'company', 'inc', 'ltd', 'corp']):
                    vendor_name = line
                    continue
                
                # Try to extract item information
                # Look for patterns like: "Item: Description - $Price x Quantity = $Total"
                if any(char in line for char in ['$', '€', '£', '¥']) and any(char in line for char in ['x', '*', '×']):
                    # This might be an item line
                    items.append({
                        'raw_text': line,
                        'description': line.split('$')[0] if '$' in line else line
                    })
            
            return {
                'vendor_name': vendor_name,
                'items': items,
                'raw_text': text
            }
            
        except Exception as e:
            print(f"Text parsing failed: {e}")
            return {
                'vendor_name': "Unknown Vendor",
                'items': [],
                'raw_text': text
            }

# Global instance
enhanced_file_processor = EnhancedFileProcessor()
